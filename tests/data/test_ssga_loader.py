"""SSGA SPY loader tests.

Synthetic fixtures matching the two vendor formats documented in
src/pit_backtest/data/sources/ssga.py:
1. Legacy CSV (distributions.csv + performance.csv).
2. SSGA-native XLSX (spdr-etf-historical-distributions.xlsx +
   spdr-product-data-us-en.xlsx).

No real SSGA data required; tests construct minimal XLSXs in-memory.
"""

from __future__ import annotations

import hashlib
import math
from datetime import date
from pathlib import Path

import openpyxl
import pytest

from pit_backtest.data.sources.manifest import SnapshotMismatchError
from pit_backtest.data.sources.ssga import (
    SSGAFormatError,
    SSGASpyReference,
    reconciliation_delta_bps,
)


_DISTRIBUTIONS_CSV = """ex_date,record_date,payable_date,amount_per_share
2023-12-15,2023-12-18,2024-01-31,1.5800
2024-03-15,2024-03-18,2024-04-30,1.7715
"""

_PERFORMANCE_CSV = """period,annualized_nav_tr_pct,annualized_market_price_tr_pct
1m,1.20,1.21
3m,5.40,5.41
ytd,8.50,8.49
1y,15.20,15.22
3y,10.10,10.11
5y,12.30,12.32
10y,11.50,11.49
si,9.95,9.94
"""


def _write_synthetic_bundle(tmp_path: Path, bundle_name: str = "spy_ssga_2026-05-28") -> Path:
    snapshots_root = tmp_path / "snapshots"
    bundle_dir = snapshots_root / bundle_name
    bundle_dir.mkdir(parents=True)

    dist_path = bundle_dir / "distributions.csv"
    dist_path.write_bytes(_DISTRIBUTIONS_CSV.encode("utf-8"))
    perf_path = bundle_dir / "performance.csv"
    perf_path.write_bytes(_PERFORMANCE_CSV.encode("utf-8"))

    dist_sha = hashlib.sha256(dist_path.read_bytes()).hexdigest()
    perf_sha = hashlib.sha256(perf_path.read_bytes()).hexdigest()
    dist_size = dist_path.stat().st_size
    perf_size = perf_path.stat().st_size

    manifest_content = f"""
[snapshots.{bundle_name}]
source = "ssga_spy"
pull_date = 2026-05-28

[snapshots.{bundle_name}.files]
"distributions.csv" = {{ sha256 = "{dist_sha}", size_bytes = {dist_size} }}
"performance.csv" = {{ sha256 = "{perf_sha}", size_bytes = {perf_size} }}
"""
    (snapshots_root / "manifest.toml").write_text(manifest_content, encoding="utf-8")
    return snapshots_root


def test_construction_verifies_manifest(tmp_path: Path) -> None:
    snapshots_root = _write_synthetic_bundle(tmp_path)
    ssga = SSGASpyReference("spy_ssga_2026-05-28", snapshots_root)
    assert ssga.bundle_name == "spy_ssga_2026-05-28"


def test_tampered_file_fails_construction(tmp_path: Path) -> None:
    snapshots_root = _write_synthetic_bundle(tmp_path)
    (snapshots_root / "spy_ssga_2026-05-28" / "distributions.csv").write_bytes(b"tampered")
    with pytest.raises(SnapshotMismatchError, match="SHA256 mismatch"):
        SSGASpyReference("spy_ssga_2026-05-28", snapshots_root)


def test_dividends_returns_sorted_frame(tmp_path: Path) -> None:
    snapshots_root = _write_synthetic_bundle(tmp_path)
    ssga = SSGASpyReference("spy_ssga_2026-05-28", snapshots_root)
    divs = ssga.dividends()
    assert divs.columns == ["ex_date", "amount_per_share"]
    assert divs["ex_date"].to_list() == [date(2023, 12, 15), date(2024, 3, 15)]
    assert divs["amount_per_share"][1] == pytest.approx(1.7715)


def test_performance_normalizes_period_casing(tmp_path: Path) -> None:
    """Period labels normalize to lowercase so lookups are case-insensitive."""
    snapshots_root = _write_synthetic_bundle(tmp_path)
    ssga = SSGASpyReference("spy_ssga_2026-05-28", snapshots_root)
    perf = ssga.performance()
    assert "10y" in perf["period"].to_list()


def test_annualized_nav_tr_lookup(tmp_path: Path) -> None:
    snapshots_root = _write_synthetic_bundle(tmp_path)
    ssga = SSGASpyReference("spy_ssga_2026-05-28", snapshots_root)
    # 10y is 11.50% in the fixture; returned as 0.1150 decimal.
    assert ssga.annualized_nav_tr_for_period("10y") == pytest.approx(0.1150)
    # Case-insensitive
    assert ssga.annualized_nav_tr_for_period("10Y") == pytest.approx(0.1150)


def test_annualized_nav_tr_unknown_period_raises(tmp_path: Path) -> None:
    snapshots_root = _write_synthetic_bundle(tmp_path)
    ssga = SSGASpyReference("spy_ssga_2026-05-28", snapshots_root)
    with pytest.raises(KeyError, match="period '20y' not in SSGA"):
        ssga.annualized_nav_tr_for_period("20y")


def test_dividends_lazy_loaded_and_cached(tmp_path: Path) -> None:
    """Repeated dividends() calls return the same frame instance (no re-read)."""
    snapshots_root = _write_synthetic_bundle(tmp_path)
    ssga = SSGASpyReference("spy_ssga_2026-05-28", snapshots_root)
    d1 = ssga.dividends()
    d2 = ssga.dividends()
    assert d1 is d2


def test_reconciliation_delta_bps_positive_when_engine_overstates() -> None:
    """Engine at 10.15%, SSGA at 10.10% = +5 bps overstate."""
    delta = reconciliation_delta_bps(0.1015, 0.1010)
    assert delta == pytest.approx(5.0, abs=1e-9)


def test_reconciliation_delta_bps_negative_when_engine_understates() -> None:
    delta = reconciliation_delta_bps(0.0995, 0.1010)
    assert delta == pytest.approx(-15.0, abs=1e-9)


def test_reconciliation_delta_bps_zero_on_match() -> None:
    delta = reconciliation_delta_bps(0.10, 0.10)
    assert delta == pytest.approx(0.0, abs=1e-12)


# ----- SSGA-native XLSX fixtures -----


def _write_distributions_xlsx(
    path: Path,
    rows: list[tuple[str, str, date, float]],
) -> None:
    """Write a minimal SSGA distributions XLSX.

    Each row is (fund_name, ticker, ex_date, dividend_amount). The
    workbook has a sheet named "dividend" with the SSGA column shape.
    """
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "dividend"
    sheet.append(
        [
            "FUND NAME",
            "TICKER",
            "CUSIP",
            "EX-DATE",
            "RECORD DATE",
            "PAYABLE DATE",
            "DIVIDEND ($)",
            "SHORT TERM CAPITAL GAIN ($)",
            "LONG TERM CAPITAL GAIN ($)",
            "FREQUENCY",
        ]
    )
    for fund_name, ticker, ex_date, dividend in rows:
        sheet.append(
            [
                fund_name,
                ticker,
                "78462F103",
                ex_date,
                ex_date,
                ex_date,
                dividend,
                0.0,
                0.0,
                "Q",
            ]
        )
    wb.save(path)


def _write_product_data_xlsx(
    path: Path,
    spy_returns: tuple[float, float, float, float, float],
    other_tickers: list[tuple[str, tuple[float, float, float, float, float]]] | None = None,
) -> None:
    """Write a minimal SSGA product-data XLSX with the two-tier header.

    spy_returns is (1y, 3y, 5y, 10y, since_inception) in percent.
    other_tickers is an optional list of additional rows to confirm the
    loader filters to SPY correctly.
    """
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Sheet1"

    # Row 1: group headers. Leave blanks under the Annualized block to
    # mirror SSGA's actual layout.
    sheet.append(
        [
            "Ticker",
            "Name",
            "Total Returns (Annualized)",
            None,
            None,
            None,
            None,
            "1 yr. FFO Growth",
        ]
    )
    # Row 2: sub-headers
    sheet.append(
        [
            None,
            None,
            "1 Year",
            "3 Year",
            "5 Year",
            "10 Year",
            "Since Inception",
            None,
        ]
    )

    if other_tickers:
        for ticker, returns in other_tickers:
            sheet.append(
                [ticker, f"{ticker} fund", *returns, 0.0]
            )

    # SPY data row
    sheet.append(
        [
            "SPY",
            "State Street SPDR S&P 500 ETF Trust",
            *spy_returns,
            0.0,
        ]
    )
    wb.save(path)


def _write_xlsx_bundle(
    tmp_path: Path,
    bundle_name: str,
    dist_rows: list[tuple[str, str, date, float]],
    spy_returns: tuple[float, float, float, float, float],
    other_tickers: list[tuple[str, tuple[float, float, float, float, float]]] | None = None,
) -> Path:
    """Write a snapshot bundle containing the two SSGA-native XLSX files
    plus a manifest with correct SHA256 entries.
    """
    snapshots_root = tmp_path / "snapshots"
    bundle_dir = snapshots_root / bundle_name
    bundle_dir.mkdir(parents=True)

    dist_path = bundle_dir / "spdr-etf-historical-distributions.xlsx"
    _write_distributions_xlsx(dist_path, dist_rows)

    prod_path = bundle_dir / "spdr-product-data-us-en.xlsx"
    _write_product_data_xlsx(prod_path, spy_returns, other_tickers)

    dist_sha = hashlib.sha256(dist_path.read_bytes()).hexdigest()
    prod_sha = hashlib.sha256(prod_path.read_bytes()).hexdigest()

    manifest_content = f"""
[snapshots.{bundle_name}]
source = "ssga_spy"
pull_date = 2026-05-29

[snapshots.{bundle_name}.files]
"spdr-etf-historical-distributions.xlsx" = {{ sha256 = "{dist_sha}", size_bytes = {dist_path.stat().st_size} }}
"spdr-product-data-us-en.xlsx" = {{ sha256 = "{prod_sha}", size_bytes = {prod_path.stat().st_size} }}
"""
    (snapshots_root / "manifest.toml").write_text(manifest_content, encoding="utf-8")
    return snapshots_root


def test_xlsx_distributions_filters_to_spy(tmp_path: Path) -> None:
    """SSGA distributions XLSX has rows for every SPDR ETF; loader keeps only SPY."""
    snapshots_root = _write_xlsx_bundle(
        tmp_path,
        bundle_name="spy_ssga_2026-05-29",
        dist_rows=[
            ("SPDR Gold Shares", "GLD", date(2024, 3, 15), 0.0),
            ("SPDR S&P 500 ETF Trust", "SPY", date(2023, 12, 15), 1.5800),
            ("SPDR S&P 500 ETF Trust", "SPY", date(2024, 3, 15), 1.7715),
            ("SPDR Bloomberg High Yield Bond ETF", "JNK", date(2024, 4, 1), 0.7),
        ],
        spy_returns=(15.2, 10.1, 12.3, 11.5, 9.95),
    )
    ssga = SSGASpyReference("spy_ssga_2026-05-29", snapshots_root)
    divs = ssga.dividends()
    assert divs.height == 2
    assert divs["ex_date"].to_list() == [date(2023, 12, 15), date(2024, 3, 15)]
    assert divs["amount_per_share"][1] == pytest.approx(1.7715)


def test_xlsx_performance_extracts_spy_annualized_returns(tmp_path: Path) -> None:
    """SSGA product-data XLSX has one row per SPDR ETF; loader extracts SPY's annualized block."""
    snapshots_root = _write_xlsx_bundle(
        tmp_path,
        bundle_name="spy_ssga_2026-05-29",
        dist_rows=[("SPDR S&P 500 ETF Trust", "SPY", date(2024, 3, 15), 1.7715)],
        spy_returns=(15.20, 10.10, 12.30, 11.50, 9.95),
        other_tickers=[
            ("GLD", (0.0, 0.0, 0.0, 0.0, 0.0)),
            ("AGG", (0.0, 0.0, 0.0, 0.0, 0.0)),
        ],
    )
    ssga = SSGASpyReference("spy_ssga_2026-05-29", snapshots_root)
    perf = ssga.performance()
    assert perf.height == 5
    assert perf["period"].to_list() == ["1y", "3y", "5y", "10y", "si"]
    assert perf["annualized_nav_tr_pct"][3] == pytest.approx(11.50)
    assert ssga.annualized_nav_tr_for_period("10y") == pytest.approx(0.1150)
    # Market price column is NaN in the XLSX path (SSGA does not expose it).
    assert math.isnan(perf["annualized_market_price_tr_pct"][0])


def test_xlsx_distributions_raises_on_missing_column(tmp_path: Path) -> None:
    """If SSGA renames the EX-DATE column, the loader raises with a clear message."""
    snapshots_root = tmp_path / "snapshots"
    bundle_dir = snapshots_root / "spy_ssga_2026-05-29"
    bundle_dir.mkdir(parents=True)

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "dividend"
    sheet.append(["FUND NAME", "TICKER", "EX_DATE_NEW_NAME", "DIVIDEND ($)"])  # wrong
    sheet.append(["SPY", "SPY", date(2024, 3, 15), 1.0])
    dist_path = bundle_dir / "spdr-etf-historical-distributions.xlsx"
    wb.save(dist_path)

    # also write a product-data XLSX so manifest verification can pass
    prod_path = bundle_dir / "spdr-product-data-us-en.xlsx"
    _write_product_data_xlsx(prod_path, (1.0, 1.0, 1.0, 1.0, 1.0))

    dist_sha = hashlib.sha256(dist_path.read_bytes()).hexdigest()
    prod_sha = hashlib.sha256(prod_path.read_bytes()).hexdigest()
    manifest = f"""
[snapshots.spy_ssga_2026-05-29]
source = "ssga_spy"
pull_date = 2026-05-29
[snapshots.spy_ssga_2026-05-29.files]
"spdr-etf-historical-distributions.xlsx" = {{ sha256 = "{dist_sha}", size_bytes = {dist_path.stat().st_size} }}
"spdr-product-data-us-en.xlsx" = {{ sha256 = "{prod_sha}", size_bytes = {prod_path.stat().st_size} }}
"""
    (snapshots_root / "manifest.toml").write_text(manifest, encoding="utf-8")

    ssga = SSGASpyReference("spy_ssga_2026-05-29", snapshots_root)
    with pytest.raises(SSGAFormatError, match="header missing"):
        ssga.dividends()


def test_xlsx_path_preferred_over_csv_when_both_present(tmp_path: Path) -> None:
    """If both XLSX and CSV are in a bundle, XLSX wins (it is SSGA's canonical export)."""
    snapshots_root = _write_xlsx_bundle(
        tmp_path,
        bundle_name="spy_ssga_2026-05-29",
        dist_rows=[("SPY", "SPY", date(2024, 3, 15), 1.7715)],
        spy_returns=(15.20, 10.10, 12.30, 11.50, 9.95),
    )
    # Also drop a CSV with conflicting data. The XLSX should win.
    bundle_dir = snapshots_root / "spy_ssga_2026-05-29"
    (bundle_dir / "distributions.csv").write_text(
        "ex_date,record_date,payable_date,amount_per_share\n"
        "2099-01-01,2099-01-01,2099-01-01,99.99\n",
        encoding="utf-8",
    )
    ssga = SSGASpyReference("spy_ssga_2026-05-29", snapshots_root)
    divs = ssga.dividends()
    # The conflicting CSV row 2099-01-01 must NOT appear; XLSX row 2024-03-15 must.
    assert date(2024, 3, 15) in divs["ex_date"].to_list()
    assert date(2099, 1, 1) not in divs["ex_date"].to_list()
