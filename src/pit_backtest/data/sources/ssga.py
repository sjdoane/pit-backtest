"""SSGA SPY reference loader (M1 reconciliation reference).

Per docs/methodology/total_return_reconstruction.md, the M1 reconciliation
compares the engine's reconstructed SPY TR to SSGA's published SPY NAV TR.
This module loads the SSGA snapshot under data/snapshots/spy_ssga_<YYYY-MM-DD>/
with the same SHA256 verification pattern as the Sharadar adapter.

The loader handles two SSGA-export shapes:

1. SSGA-native XLSX (preferred, what SSGA actually publishes today):
   - spdr-etf-historical-distributions.xlsx (sheet "dividend"): per-distribution
     rows for every SPDR ETF; the loader filters to TICKER='SPY' and maps
     EX-DATE, DIVIDEND ($), to ex_date, amount_per_share.
   - spdr-product-data-us-en.xlsx (sheet "Sheet1"): wide one-row-per-ETF
     snapshot with a two-tier column header. The loader uses openpyxl
     directly to walk the headers, find SPY's row, and extract the
     "Total Returns (Annualized)" cells for 1 Year / 3 Year / 5 Year /
     10 Year / Since Inception.

2. Legacy CSV (synthetic tests and any prior manual exports):
   - distributions.csv: ex_date, record_date, payable_date, amount_per_share
   - performance.csv: period, annualized_nav_tr_pct, annualized_market_price_tr_pct

The XLSX format wins when both are present in a bundle.

Vendor URL of record (verified 2026-05-29):
https://www.ssga.com/us/en/intermediary/etfs/spdr-sp-500-etf-spy

SSGA's older URL (https://www.ssga.com/us/en/intermediary/etfs/spy) no
longer redirects. See docs/vendor/nasdaq-data-link-pull.md for the
download workflow.
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any

import polars as pl

from pit_backtest.data.sources.manifest import load_manifest, verify_bundle


# Legacy CSV filenames (kept for synthetic tests and any prior manual exports).
_DISTRIBUTIONS_CSV_FILENAME = "distributions.csv"
_PERFORMANCE_CSV_FILENAME = "performance.csv"

# SSGA-native XLSX filenames as the vendor publishes them today.
_DISTRIBUTIONS_XLSX_FILENAME = "spdr-etf-historical-distributions.xlsx"
_PRODUCT_DATA_XLSX_FILENAME = "spdr-product-data-us-en.xlsx"
_PRODUCT_DATA_XLSX_SHEET = "Sheet1"
_DISTRIBUTIONS_XLSX_SHEET = "dividend"


# SSGA column labels in spdr-etf-historical-distributions.xlsx. Capture the
# spellings here so any future renames surface in one place.
_DIST_TICKER_COL = "TICKER"
_DIST_EX_DATE_COL = "EX-DATE"
_DIST_AMOUNT_COL = "DIVIDEND ($)"

# SSGA column labels in spdr-product-data-us-en.xlsx. The product data
# workbook has a two-tier header: row 1 carries group names, row 2 carries
# the period sub-labels. We anchor on the row-1 group name "Total Returns
# (Annualized)" and then read the next 5 columns whose row-2 sub-labels
# are "1 Year", "3 Year", "5 Year", "10 Year", "Since Inception".
_PROD_TICKER_HEADER = "Ticker"
_PROD_ANN_RETURNS_HEADER = "Total Returns (Annualized)"
_PROD_AS_OF_HEADER = "Total Returns as of Date"
_PROD_ANN_PERIOD_LABELS = ("1 Year", "3 Year", "5 Year", "10 Year", "Since Inception")
# Map SSGA's "1 Year" -> our canonical "1y" tag used everywhere downstream.
_SSGA_PERIOD_TO_TAG = {
    "1 Year": "1y",
    "3 Year": "3y",
    "5 Year": "5y",
    "10 Year": "10y",
    "Since Inception": "si",
}


class SSGAFormatError(ValueError):
    """Raised when the SSGA XLSX schema does not match the expected shape.

    Typical cause: SSGA renamed a column header or changed the row layout.
    Diagnose by opening the XLSX in Excel and comparing to the constants
    at the top of this module.
    """


class SSGASpyReference:
    """Loads the SSGA SPY snapshot for the M1 reconciliation."""

    def __init__(self, snapshot_bundle: str, snapshots_root: Path) -> None:
        self._bundle_name = snapshot_bundle
        self._snapshots_root = snapshots_root.resolve()
        manifest_path = self._snapshots_root / "manifest.toml"
        self._manifest = load_manifest(manifest_path)
        verify_bundle(snapshot_bundle, self._snapshots_root, self._manifest)
        self._bundle_dir = self._snapshots_root / snapshot_bundle

        self._distributions: pl.DataFrame | None = None
        self._performance: pl.DataFrame | None = None
        self._as_of_date: date | None = None

    @property
    def bundle_name(self) -> str:
        return self._bundle_name

    @property
    def as_of_date(self) -> date | None:
        """The as-of date for SSGA's published trailing returns.

        SSGA's product-data returns are trailing periods ending at this
        date (e.g., "10y" is the 10 years ending here). The reconciliation
        harness aligns the engine's window to this date per ADR 0006.
        None for the legacy CSV path, which had no as-of column. Triggers
        a performance() load if not yet read.
        """
        if self._performance is None:
            self.performance()
        return self._as_of_date

    def dividends(self) -> pl.DataFrame:
        """Return the SSGA-published SPY distribution history.

        Columns: ex_date (pl.Date), amount_per_share (pl.Float64).
        Sorted by ex_date for determinism.

        Reads spdr-etf-historical-distributions.xlsx if present (filtered
        to TICKER='SPY'); otherwise falls back to distributions.csv.
        """
        if self._distributions is not None:
            return self._distributions

        xlsx_path = self._bundle_dir / _DISTRIBUTIONS_XLSX_FILENAME
        csv_path = self._bundle_dir / _DISTRIBUTIONS_CSV_FILENAME
        if xlsx_path.is_file():
            self._distributions = _read_distributions_xlsx(xlsx_path)
        elif csv_path.is_file():
            raw = pl.read_csv(csv_path, try_parse_dates=True)
            self._distributions = raw.select(
                pl.col("ex_date").cast(pl.Date),
                pl.col("amount_per_share").cast(pl.Float64),
            ).sort("ex_date")
        else:
            raise FileNotFoundError(
                f"SSGA bundle has neither {_DISTRIBUTIONS_XLSX_FILENAME} nor "
                f"{_DISTRIBUTIONS_CSV_FILENAME} under {self._bundle_dir}"
            )
        return self._distributions

    def performance(self) -> pl.DataFrame:
        """Return the SSGA-published SPY performance summary.

        Columns: period (pl.String), annualized_nav_tr_pct (pl.Float64),
        annualized_market_price_tr_pct (pl.Float64).

        Reads spdr-product-data-us-en.xlsx if present (extracts SPY's row
        and the Annualized Total Returns block); otherwise falls back to
        performance.csv.

        Note: the XLSX path populates annualized_market_price_tr_pct with
        NaN because SSGA's product-data workbook reports only NAV-based
        annualized returns by default. The reconciliation uses
        annualized_nav_tr_pct exclusively, so this is acceptable. The CSV
        path requires both columns to be present per the legacy schema.
        """
        if self._performance is not None:
            return self._performance

        xlsx_path = self._bundle_dir / _PRODUCT_DATA_XLSX_FILENAME
        csv_path = self._bundle_dir / _PERFORMANCE_CSV_FILENAME
        if xlsx_path.is_file():
            self._performance, self._as_of_date = read_performance_xlsx_with_as_of(
                xlsx_path
            )
        elif csv_path.is_file():
            raw = pl.read_csv(csv_path)
            self._performance = raw.select(
                pl.col("period").cast(pl.String).str.to_lowercase(),
                pl.col("annualized_nav_tr_pct").cast(pl.Float64),
                pl.col("annualized_market_price_tr_pct").cast(pl.Float64),
            )
        else:
            raise FileNotFoundError(
                f"SSGA bundle has neither {_PRODUCT_DATA_XLSX_FILENAME} nor "
                f"{_PERFORMANCE_CSV_FILENAME} under {self._bundle_dir}"
            )
        return self._performance

    def annualized_nav_tr_for_period(self, period: str) -> float:
        """Return SSGA's published annualized NAV TR for a labeled period.

        period is one of '1m', '3m', 'ytd', '1y', '3y', '5y', '10y', 'si'.
        Returned as a decimal (e.g., 0.1234 for 12.34%/yr), not percent.
        """
        normalized = period.lower()
        perf = self.performance().filter(pl.col("period") == normalized)
        if perf.height == 0:
            raise KeyError(
                f"period '{period}' not in SSGA performance snapshot; "
                f"available: {sorted(self.performance()['period'].to_list())}"
            )
        if perf.height > 1:
            raise ValueError(
                f"period '{period}' is duplicated in SSGA performance snapshot; "
                f"manifest may be stale"
            )
        return float(perf["annualized_nav_tr_pct"][0]) / 100.0


def reconciliation_delta_bps(
    engine_annualized_return: float, ssga_annualized_return: float
) -> float:
    """Return the engine-vs-SSGA annualized return delta in basis points.

    Positive = engine overstates relative to SSGA. The M1 kill-early gate
    asserts abs(delta_bps) <= 5 bps over the 2005-2024 window.
    """
    return (engine_annualized_return - ssga_annualized_return) * 10_000.0


def _clean_ticker(raw: object) -> str:
    """Normalize an SSGA ticker cell to its bare symbol.

    SSGA decorates tickers with a registered-trademark suffix (SPY becomes
    "SPY" + U+00AE), which surfaces as "SPYM" / "SPY?" depending on console
    encoding. Take the leading run of [A-Za-z0-9.] so "SPY(R)" -> "SPY",
    "GLD(R)" -> "GLD", "GLDM(R)" -> "GLDM" (distinct funds stay distinct).
    """
    import re

    match = re.match(r"[A-Za-z0-9.]+", str(raw).strip())
    return match.group(0) if match else ""


def _parse_ssga_date(raw: object) -> date:
    """Parse an SSGA date cell.

    The distributions XLSX stores ex-dates as text "MM/DD/YYYY"; openpyxl
    returns the string. Some cells may already be datetime if SSGA changes
    the formatting, so handle both.
    """
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    text = str(raw).strip()
    return datetime.strptime(text, "%m/%d/%Y").date()


def _parse_ssga_pct(raw: object) -> float | None:
    """Parse an SSGA percent cell to a float in percent.

    Product-data returns are text like "31.01%"; strip the trailing percent
    sign and surrounding whitespace. SSGA uses "-" for not-applicable;
    return None for those. The dividend column is a bare number-as-text
    ("1.796999" or " 0.859222 "); the same parser handles it (no percent
    sign to strip).
    """
    if raw is None:
        return None
    text = str(raw).strip().rstrip("%").strip()
    if text in ("", "-"):
        return None
    return float(text)


def _find_header_row(
    all_rows: list[tuple[Any, ...]], marker: str, max_scan: int = 12
) -> int:
    """Return the index of the first row containing `marker` as a cell value.

    SSGA prepends a disclaimer paragraph row (and sometimes blanks) before
    the real header row, so the header is not always row 0. Scan the first
    `max_scan` rows. Raises SSGAFormatError if not found.
    """
    for idx, row in enumerate(all_rows[:max_scan]):
        if row is not None and marker in row:
            return idx
    raise SSGAFormatError(
        f"could not find header row containing {marker!r} in the first "
        f"{max_scan} rows of the SSGA XLSX; the export layout may have changed"
    )


def _read_distributions_xlsx(path: Path) -> pl.DataFrame:
    """Read SSGA's distributions XLSX and produce our canonical schema.

    Filters to TICKER='SPY'. Returns columns (ex_date, amount_per_share)
    sorted by ex_date. Handles MM/DD/YYYY date strings and whitespace-
    padded number-as-text dividend cells.
    """
    import openpyxl  # type: ignore[import-untyped]

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if _DISTRIBUTIONS_XLSX_SHEET not in workbook.sheetnames:
        raise SSGAFormatError(
            f"SSGA distributions XLSX missing sheet '{_DISTRIBUTIONS_XLSX_SHEET}'; "
            f"found {workbook.sheetnames}"
        )
    sheet = workbook[_DISTRIBUTIONS_XLSX_SHEET]
    all_rows = list(sheet.iter_rows(values_only=True))
    workbook.close()

    header_idx = _find_header_row(all_rows, _DIST_TICKER_COL)
    header = list(all_rows[header_idx])
    try:
        ticker_col = header.index(_DIST_TICKER_COL)
        ex_date_col = header.index(_DIST_EX_DATE_COL)
        amount_col = header.index(_DIST_AMOUNT_COL)
    except ValueError as e:
        raise SSGAFormatError(
            f"SSGA distributions XLSX header missing one of "
            f"({_DIST_TICKER_COL}, {_DIST_EX_DATE_COL}, {_DIST_AMOUNT_COL}); "
            f"got headers {header}"
        ) from e

    ex_dates: list[date] = []
    amounts: list[float] = []
    for row in all_rows[header_idx + 1 :]:
        if row is None:
            continue
        raw_ticker = row[ticker_col] if len(row) > ticker_col else None
        if raw_ticker is None or _clean_ticker(raw_ticker) != "SPY":
            continue
        ex_raw = row[ex_date_col]
        amt = _parse_ssga_pct(row[amount_col])
        if ex_raw is None or amt is None:
            continue
        ex_dates.append(_parse_ssga_date(ex_raw))
        amounts.append(amt)

    if not ex_dates:
        raise SSGAFormatError(
            f"SSGA distributions XLSX has no SPY rows; check TICKER column "
            f"content. File: {path}"
        )

    return (
        pl.DataFrame({"ex_date": ex_dates, "amount_per_share": amounts})
        .with_columns(pl.col("ex_date").cast(pl.Date))
        .sort("ex_date")
    )


def read_performance_xlsx_with_as_of(path: Path) -> tuple[pl.DataFrame, date | None]:
    """Read SSGA's product-data XLSX; return (performance_frame, as_of_date).

    Returns columns (period, annualized_nav_tr_pct, annualized_market_price_tr_pct).
    annualized_market_price_tr_pct is NaN because the workbook does not
    expose a Market Price annualized return; the reconciliation uses the
    NAV column exclusively.

    The returns are TRAILING periods ending at as_of_date (e.g., "10y" is
    the 10 years ending at as_of_date, NOT a fixed 2005-2024 window). The
    reconciliation harness must align the engine's window to as_of_date;
    see ADR 0006.
    """
    import math

    import openpyxl

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if _PRODUCT_DATA_XLSX_SHEET not in workbook.sheetnames:
        raise SSGAFormatError(
            f"SSGA product-data XLSX missing sheet '{_PRODUCT_DATA_XLSX_SHEET}'; "
            f"found {workbook.sheetnames}"
        )
    sheet = workbook[_PRODUCT_DATA_XLSX_SHEET]
    all_rows = list(sheet.iter_rows(values_only=True))
    workbook.close()

    # SSGA prepends a disclaimer row; the group-header row is the one that
    # contains "Ticker", and the period sub-labels are the row after it.
    header_idx = _find_header_row(all_rows, _PROD_TICKER_HEADER)
    header_row = list(all_rows[header_idx])
    subheader_row = list(all_rows[header_idx + 1])

    ticker_col = header_row.index(_PROD_TICKER_HEADER)
    try:
        ann_block_start = header_row.index(_PROD_ANN_RETURNS_HEADER)
    except ValueError as e:
        raise SSGAFormatError(
            f"SSGA product-data XLSX header row missing "
            f"'{_PROD_ANN_RETURNS_HEADER}'; got {header_row}"
        ) from e

    as_of_col = (
        header_row.index(_PROD_AS_OF_HEADER)
        if _PROD_AS_OF_HEADER in header_row
        else None
    )

    period_cols: dict[str, int] = {}
    for offset, expected_label in enumerate(_PROD_ANN_PERIOD_LABELS):
        col = ann_block_start + offset
        actual_label = subheader_row[col] if col < len(subheader_row) else None
        if actual_label != expected_label:
            raise SSGAFormatError(
                f"SSGA product-data XLSX sub-header column {col} expected "
                f"'{expected_label}' but got '{actual_label}'. The Annualized "
                f"Total Returns block layout has changed; update _PROD_ANN_PERIOD_LABELS."
            )
        period_cols[_SSGA_PERIOD_TO_TAG[expected_label]] = col

    spy_row: tuple[Any, ...] | None = None
    for row in all_rows[header_idx + 1 :]:
        if row is None:
            continue
        if len(row) > ticker_col and _clean_ticker(row[ticker_col]) == "SPY":
            spy_row = row
            break

    if spy_row is None:
        raise SSGAFormatError(
            f"SSGA product-data XLSX has no SPY row in column {ticker_col}; "
            f"file may be missing the SPY entry or the workbook structure has changed."
        )

    as_of_date: date | None = None
    if as_of_col is not None and len(spy_row) > as_of_col:
        as_of_raw = spy_row[as_of_col]
        if as_of_raw is not None:
            as_of_date = _parse_as_of_date(as_of_raw)

    periods: list[str] = []
    nav_pct: list[float] = []
    mkt_pct: list[float] = []
    for tag in ("1y", "3y", "5y", "10y", "si"):
        col = period_cols[tag]
        parsed = _parse_ssga_pct(spy_row[col] if col < len(spy_row) else None)
        if parsed is None:
            continue
        periods.append(tag)
        nav_pct.append(parsed)
        mkt_pct.append(math.nan)

    frame = pl.DataFrame(
        {
            "period": periods,
            "annualized_nav_tr_pct": nav_pct,
            "annualized_market_price_tr_pct": mkt_pct,
        }
    )
    return frame, as_of_date


def _parse_as_of_date(raw: object) -> date | None:
    """Parse SSGA's "Total Returns as of Date" cell, e.g. "Apr 30 2026"."""
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    text = str(raw).strip()
    for fmt in ("%b %d %Y", "%B %d %Y", "%m/%d/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _read_performance_xlsx(path: Path) -> pl.DataFrame:
    """Backwards-compatible wrapper returning just the performance frame."""
    frame, _as_of = read_performance_xlsx_with_as_of(path)
    return frame
